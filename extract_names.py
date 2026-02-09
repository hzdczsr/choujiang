import openpyxl
import json

try:
    # 打开Excel文件
    workbook = openpyxl.load_workbook('d:\\原C盘\\download\\choujiang\\namexlsx.xlsx')
    
    print(f"工作表名称: {workbook.sheetnames}")
    
    names = []
    
    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:
        print(f"\n处理工作表: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # 获取最大行数和列数
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        print(f"最大行数: {max_row}, 最大列数: {max_col}")
        
        # 遍历所有单元格
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    # 清理单元格值
                    cell_str = str(cell_value).strip()
                    if cell_str and cell_str != 'nan' and len(cell_str) > 0:
                        names.append(cell_str)
    
    # 去重
    names = list(set(names))
    
    print(f"\n提取到的姓名（共{len(names)}个）:")
    for i, name in enumerate(sorted(names), 1):
        print(f"{i}. {name}")
    
    # 保存到JSON文件
    with open('names.json', 'w', encoding='utf-8') as f:
        json.dump(sorted(names), f, ensure_ascii=False, indent=2)
    
    print(f"\n姓名数据已保存到 names.json")
    
except Exception as e:
    print(f"处理Excel文件时出错: {e}")
    import traceback
    traceback.print_exc()