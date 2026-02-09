#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl
import json
import os
from datetime import datetime

app = Flask(__name__)
CORS(app)

# 数据存储路径
DATA_DIR = 'data'
os.makedirs(DATA_DIR, exist_ok=True)

@app.route('/')
def index():
    return send_from_directory('.', 'lottery.html')

@app.route('/api/upload-excel', methods=['POST'])
def upload_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有文件上传'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': '只支持Excel文件格式'}), 400
        
        # 保存文件
        filename = f"uploaded_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(DATA_DIR, filename)
        file.save(filepath)
        
        # 读取Excel文件
        try:
            workbook = openpyxl.load_workbook(filepath)
            sheet_names = workbook.sheetnames
            print(f"工作表名称: {sheet_names}")
            
            all_names = []
            sheet_data = {}
            
            # 遍历所有工作表
            for sheet_name in sheet_names:
                print(f"处理工作表: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # 获取最大行数和列数
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                print(f"最大行数: {max_row}, 最大列数: {max_col}")
                
                sheet_names_list = []
                
                # 遍历所有单元格
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value:
                            # 清理单元格值
                            cell_str = str(cell_value).strip()
                            if cell_str and cell_str != 'nan' and len(cell_str) > 0:
                                sheet_names_list.append(cell_str)
                
                # 去重
                sheet_names_list = list(set(sheet_names_list))
                sheet_data[sheet_name] = sheet_names_list
                all_names.extend(sheet_names_list)
            
            # 整体去重
            all_names = list(set(all_names))
            
            # 保存数据到JSON文件
            data = {
                'filename': filename,
                'original_filename': file.filename,
                'upload_time': datetime.now().isoformat(),
                'sheet_data': sheet_data,
                'all_names': sorted(all_names),
                'total_count': len(all_names)
            }
            
            json_filename = filename.replace('.xlsx', '.json')
            json_filepath = os.path.join(DATA_DIR, json_filename)
            
            with open(json_filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            print(f"提取到的姓名（共{len(all_names)}个）:")
            for i, name in enumerate(sorted(all_names), 1):
                print(f"{i}. {name}")
            
            return jsonify({
                'success': True,
                'data': data
            })
            
        except Exception as e:
            return jsonify({'error': f'读取Excel文件失败: {str(e)}'}), 500
            
    except Exception as e:
        return jsonify({'error': f'处理文件失败: {str(e)}'}), 500

@app.route('/api/get-names', methods=['GET'])
def get_names():
    try:
        # 获取最新的数据文件
        json_files = [f for f in os.listdir(DATA_DIR) if f.endswith('.json')]
        if not json_files:
            return jsonify({'error': '没有找到数据文件'}), 404
        
        # 按修改时间排序，获取最新的
        json_files.sort(key=lambda x: os.path.getmtime(os.path.join(DATA_DIR, x)), reverse=True)
        latest_file = json_files[0]
        
        with open(os.path.join(DATA_DIR, latest_file), 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        return jsonify({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return jsonify({'error': f'读取数据失败: {str(e)}'}), 500

@app.route('/api/delete-data', methods=['DELETE'])
def delete_data():
    try:
        # 删除所有数据文件
        json_files = [f for f in os.listdir(DATA_DIR) if f.endswith('.json')]
        excel_files = [f for f in os.listdir(DATA_DIR) if f.endswith(('.xlsx', '.xls'))]
        
        for file in json_files + excel_files:
            os.remove(os.path.join(DATA_DIR, file))
        
        return jsonify({
            'success': True,
            'message': '数据已清除'
        })
        
    except Exception as e:
        return jsonify({'error': f'删除数据失败: {str(e)}'}), 500

@app.route('/api/get-files', methods=['GET'])
def get_files():
    try:
        json_files = [f for f in os.listdir(DATA_DIR) if f.endswith('.json')]
        excel_files = [f for f in os.listdir(DATA_DIR) if f.endswith(('.xlsx', '.xls'))]
        
        files_info = []
        
        # 处理JSON文件
        for json_file in json_files:
            filepath = os.path.join(DATA_DIR, json_file)
            stat = os.stat(filepath)
            
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            files_info.append({
                'filename': data.get('original_filename', json_file),
                'upload_time': data.get('upload_time', ''),
                'total_count': data.get('total_count', 0),
                'sheet_count': len(data.get('sheet_data', {})),
                'type': 'data'
            })
        
        # 按上传时间排序
        files_info.sort(key=lambda x: x['upload_time'], reverse=True)
        
        return jsonify({
            'success': True,
            'files': files_info
        })
        
    except Exception as e:
        return jsonify({'error': f'获取文件列表失败: {str(e)}'}), 500

if __name__ == '__main__':
    print("启动Excel抽奖服务器...")
    print("访问地址: http://localhost:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)