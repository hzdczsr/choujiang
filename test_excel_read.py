#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import json

def test_excel_reading():
    """测试读取Excel文件功能"""
    excel_file = 'namexlsx.xlsx'
    
    try:
        print(f"正在读取Excel文件: {excel_file}")
        
        # 打开Excel文件
        workbook = openpyxl.load_workbook(excel_file)
        sheet_names = workbook.sheetnames
        
        print(f"工作表数量: {len(sheet_names)}")
        print(f"工作表名称: {sheet_names}")
        
        all_names = []
        sheet_data = {}
        
        # 遍历所有工作表
        for sheet_name in sheet_names:
            print(f"\n处理工作表: {sheet_name}")
            sheet = workbook[sheet_name]
            
            # 获取最大行数和列数
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            print(f"最大行数: {max_row}, 最大列数: {max_col}")
            
            sheet_names_list = []
            
            # 遍历所有单元格
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value:
                        # 清理单元格值
                        cell_str = str(cell_value).strip()
                        if cell_str and cell_str != 'nan' and len(cell_str) > 0:
                            sheet_names_list.append(cell_str)
                            all_names.append(cell_str)
            
            # 去重
            sheet_names_list = list(set(sheet_names_list))
            sheet_data[sheet_name] = sorted(sheet_names_list)
            
            print(f"工作表 '{sheet_name}' 中的姓名数量: {len(sheet_names_list)}")
            print(f"前10个姓名: {sheet_names_list[:10]}")
        
        # 整体去重
        all_names = list(set(all_names))
        all_names.sort()
        
        print(f"\n总共提取到的姓名数量: {len(all_names)}")
        print(f"所有姓名: {all_names}")
        
        # 保存到JSON文件
        data = {
            'filename': excel_file,
            'sheet_names': sheet_names,
            'sheet_data': sheet_data,
            'all_names': all_names,
            'total_count': len(all_names)
        }
        
        with open('extracted_names.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"\n姓名数据已保存到 extracted_names.json")
        print("Excel文件读取测试完成！")
        
        return True
        
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_excel_reading()